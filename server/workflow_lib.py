import os
import io
import json
import fitz  # PyMuPDF

from google import genai
from google.genai import types
from thefuzz import fuzz
import jellyfish
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment
from copy import copy
fitz.TOOLS.mupdf_display_errors(False)

# Load environment variables from .env file if it exists
load_dotenv()

# CONFIGURATION
API_KEY = os.getenv("GEMINI_API_KEY")

if not API_KEY:
    # Fallback to an empty string to avoid crashes, but warn the user
    print("WARNING: GEMINI_API_KEY not found in environment or .env file.")
    API_KEY = "MISSING_KEY"

MODEL_NAME = "gemini-3-flash-preview"

# Column mapping for Excel: col_index -> JSON key (None = special handling)
COL_MAP = {
    1: None,                  # S/NO (auto-increment)
    2: "Project Name",
    3: "Constituency",
    4: "County",
    5: None,                  # Region (empty)
    6: None,                  # Affected land (empty)
    7: "Plot No",
    8: "Owned by",
    9: "Signed by",
    10: "Relationship",
    11: "ID No",
    12: "Phone No",
    13: "Ownership Document",
    14: "Consent Signed",
}

class ConsentExtractor:
    def __init__(self, model_name=MODEL_NAME):
        self.client = genai.Client(api_key=API_KEY)
        self.model_name = model_name

    def extract_details(self, pdf_path, processed_pages=None, global_offset=0, global_total=None):
        """
        Extracts details from ALL pages of the PDF in parallel.
        Yields: (page_num, event_dict)
        """
        if processed_pages is None:
            processed_pages = set()
            
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        doc = None
        try:
            doc = fitz.open(pdf_path)
            num_pages = len(doc)
            target_total = global_total if global_total is not None else num_pages
            
            print(f"  Analysing {num_pages} pages for {os.path.basename(pdf_path)}...")
            yield 0, {"type": "progress", "current": global_offset, "total": target_total, "status": f"Initializing {os.path.basename(pdf_path)}..."}

            results = [None] * num_pages
            completed_count = 0
            chunk_size = 20 # Process 20 pages at a time to keep RAM flat
            
            for chunk_start in range(0, num_pages, chunk_size):
                chunk_end = min(chunk_start + chunk_size, num_pages)
                p_tasks = []
                
                for page_num in range(chunk_start, chunk_end):
                    if page_num in processed_pages:
                        completed_count += 1
                        continue
                    
                    # Store task info, rendering will happen in parallel per-thread
                    p_tasks.append(page_num)
                
                if not p_tasks:
                    continue

                with ThreadPoolExecutor(max_workers=10) as executor:
                    future_to_page = {
                        executor.submit(self.process_page_parallel, pdf_path, p_num): p_num 
                        for p_num in p_tasks
                    }
                    
                    for future in as_completed(future_to_page):
                        p_num = future_to_page[future]
                        completed_count += 1
                        data = future.result()
                        results[p_num] = data
                        
                        abs_current = global_offset + completed_count
                        yield p_num, {"type": "progress", "current": abs_current, "total": target_total, "status": f"Scanning page {p_num + 1}..."}

            # Sequential Pass to resolve "Ownership Document"
            print(f"  Finalizing {num_pages} pages for {os.path.basename(pdf_path)}...")
            for i in range(num_pages):
                data = results[i]
                if data and data.get("is_wayleave_consent_form"):
                    print(f"    [Page {i+1}] CONSENT FORM Found. Resolving ownership...")
                    # Scan forward until the next consent form or EOF
                    ownership_doc = "UNDER ADJUDICATION"
                    for j in range(i + 1, num_pages):
                        next_data = results[j]
                        if not next_data: continue
                        
                        if next_data.get("is_wayleave_consent_form"):
                            break # Hit the next person's form
                        
                        dt = next_data.get("document_type", "OTHER")
                        if dt == "TITLE_DEED":
                            print(f"      - Matched TITLE DEED (Page {j+1})")
                            ownership_doc = "TITLE DEED"
                            break # Conclusive for this group
                        elif dt == "LAND_SALE_AGREEMENT":
                            if ownership_doc in ["UNDER ADJUDICATION", "SEARCH AWAITED"]:
                                print(f"      - Matched LAND SALE AGREEMENT (Page {j+1})")
                                ownership_doc = "LAND SALE AGREEMENT"
                        elif dt == "SEARCH_DOCUMENT":
                            if ownership_doc == "UNDER ADJUDICATION":
                                print(f"      - Matched SEARCH DOCUMENT (Page {j+1})")
                                ownership_doc = "SEARCH AWAITED"
                    
                    data["Ownership Document"] = ownership_doc
                    yield i, {"type": "data", "page_num": i, "data": data}
                elif data:
                    dtype = data.get('document_type', 'OTHER')
                    print(f"    [Page {i+1}] Identified as: {dtype}")
                    yield i, {"type": "skip", "page_num": i}
                else:
                    yield i, {"type": "skip", "page_num": i}
                    
        except Exception as e:
            print(f"[Extractor Fatal Error] {e}")
            raise e # Actively raise to the caller (main.py) so it can send an error stream event
        finally:
            if doc:
                doc.close()
                print(f"  Closed PDF handle for {pdf_path}")

    def process_page_parallel(self, pdf_path, page_num):
        """Helper for parallel execution: Handles rendering and AI call in one thread."""
        try:
            # We open a local handle in each thread for total safety
            doc = fitz.open(pdf_path)
            page = doc[page_num]
            pix = page.get_pixmap(dpi=150)
            img_data = pix.tobytes("png")
            
            # Store metadata needed for snippet creation later
            page_width = page.rect.width
            page_height = page.rect.height
            rotation = page.rotation
            doc.close()

            data = self.execute_gemini_request(img_data, page_num)
            if data:
                data["page_width"] = page_width
                data["page_height"] = page_height
                data["rotation"] = rotation
            return data
        except Exception as e:
            print(f"    [Thread Error Page {page_num+1}] {e}")
            return None

    def execute_gemini_request(self, img_data, page_num):
        """Core Gemini call logic separated for parallel use"""
        try:
            prompt = """
            Identify the type of this document and extract details if it is a 'WAYLEAVE CONSENT FORM'.
            
            Document Types:
            - 'WAYLEAVE_CONSENT_FORM': Standard form with 'WAYLEAVE CONSENT FORM' title and fields for Land owner, Plot No, etc.
            - 'TITLE_DEED': Land ownership document (e.g., Certificate of Title, Lease, Allotment Letter).
            - 'LAND_SALE_AGREEMENT': Agreement documenting land transfer/sale between parties.
            - 'SEARCH_DOCUMENT': Official land search results/documents from the land registry.
            - 'ID_PHOTOCOPY': Photocopy of a National ID or Passport.
            - 'OTHER': Any other document.

            Return ONLY valid JSON -- no markdown, no explanation.

            {
              "document_type": "WAYLEAVE_CONSENT_FORM" | "TITLE_DEED" | "LAND_SALE_AGREEMENT" | "SEARCH_DOCUMENT" | "ID_PHOTOCOPY" | "OTHER",
              "is_wayleave_consent_form": true or false,
              "Project Name": "Name of project/village (only for CONSENT FORM)",
              "Constituency": "Constituency name (only for CONSENT FORM)",
              "County": "County name (only for CONSENT FORM)",
              "Plot No": "EXTRACT ONLY THE DIGITS (only for CONSENT FORM)",
              "Owned by": "Full name of the land owner (only for CONSENT FORM)",
              "Signed by": "Full name of signer (only for CONSENT FORM)",
              "Relationship": "Relationship to owner (only for CONSENT FORM)",
              "ID No": "ID number (only for CONSENT FORM)",
              "Phone No": "Phone number (only for CONSENT FORM)",
              "Consent Signed": "YES or NO (only for CONSENT FORM)",
              "proprietor_name": "Same as 'Signed by'",
              "title_number": "Same as 'Plot No'",
              "sketch_box_1000": [ymin, xmin, ymax, xmax]
            }

            Rules:
            - For 'Plot No', extract ONLY the numerical digits.
            - Handwritten Signatures: DO NOT transcribe cursive or handwritten signatures as text for names. Look for a PRINTED/TYPED version of the name. If only a signature is present without a nearby printed name in the field, return an empty string (""). NEVER attempt to "spell out" a scribble.
            - If not a WAYLEAVE_CONSENT_FORM, return null for all fields except document_type and is_wayleave_consent_form.
            """
            
            import time
            retries = 5
            response = None
            for attempt in range(retries):
                try:
                    response = self.client.models.generate_content(
                        model=self.model_name,
                        contents=[
                            prompt,
                            types.Part.from_bytes(data=img_data, mime_type="image/png")
                        ]
                    )
                    break 
                except Exception as e:
                    err_msg = str(e).upper()
                    # Broader error matching for networking issues
                    is_network_error = any(x in err_msg for x in [
                        "503", "UNAVAILABLE", "RATE_LIMIT", "QUOTA", 
                        "10053", "10054", "11001", "SSL", "EOF", 
                        "CONNECTION", "ABORTED", "RESET", "GETADDRINFO"
                    ])
                    
                    if is_network_error:
                        if attempt < retries - 1:
                            # Exponential backoff: 2s, 4s, 8s, 16s...
                            wait_time = 2 ** (attempt + 1)
                            print(f"    [Retry {attempt+1}] Networking/API issue on Page {page_num+1}: {e}. Backing off {wait_time}s...")
                            time.sleep(wait_time)
                            continue
                    
                    print(f"    [Page {page_num+1} Final Failure] {e}")
                    raise Exception(f"Fatal API Error on page {page_num+1} after {retries} retries: {str(e)}")

            if not response or not response.text:
                return None
                
            text = response.text.strip()
            if "```json" in text:
                text = text.split("```json")[-1].split("```")[0].strip()
            elif "```" in text:
                text = text.split("```")[-1].split("```")[0].strip()
            
            try:
                data = json.loads(text)
            except json.JSONDecodeError as je:
                print(f"    [Page {page_num+1} Error] Invalid JSON from AI: {je}\nResponse text: {text}")
                return None
            
            return data
        except Exception as e:
            # Re-raise explicit exceptions (like the retry limit exception above)
            if "Fatal API Error" in str(e):
                raise e
            print(f"    [Page {page_num+1} Exception] {e}")
            return None

    def process_page(self, doc, page_num):
        """
        Extracts details from a specific page.
        Returns: data_json or None if failed.
        """
        try:
            page = doc[page_num]
            pix = page.get_pixmap(dpi=150) # 150 DPI is enough for text reading
            img_data = pix.tobytes("png")
            
            data = self.execute_gemini_request(img_data, page_num)
            
            if data:
                # Add PDF page dimensions for coordinate conversion
                data["page_width"] = page.rect.width
                data["page_height"] = page.rect.height
                data["rotation"] = page.rotation
            
            return data
        except Exception as e:
            print(f"    [Page {page_num+1} Error] {e}")
            return None

class SitePlanLocator:
    def __init__(self, site_plan_path):
        self.path = site_plan_path
        self.doc = fitz.open(site_plan_path)
        self.index = [] # List of {text, page, rect}
        self._build_index()

    def _build_index(self):
        print("Indexing Site Plan (this may take a moment)...")
        # Optimization: We currently use direct fitz search, no indexing needed for now.
        pass

    def close(self):
        """Explicitly close the PDF document."""
        if hasattr(self, 'doc') and self.doc:
            try:
                self.doc.close()
            except Exception:
                pass
    
    def search(self, name, title_number):
        """
        Search for Proprietor Name (Fuzzy) or Plot Number (Exact/Vicinity).
        name: The main Proprietor name to search for.
        Returns: {page, rect, method} or None.
        """
        import re
        best_match = None
        best_score = 0
        
        # 1. Search by Name (Hybrid TF-IDF N-Gram + Phonetic)
        if name:
            print(f"  Searching for Name '{name}'...", end="", flush=True)
            
            # Words > 2 chars to avoid initials/noise
            name_parts = [p for p in name.lower().split() if len(p) > 2]
            if not name_parts: return None
            
            clean_target_name = " ".join(name_parts)
            
            # Pre-calculate phonetic codes for the target name parts
            # Clean non-alphabetical chars as jellyfish match_rating_codex strictly requires them
            target_phonetics = [jellyfish.match_rating_codex(re.sub(r'[^a-zA-Z]', '', p)) for p in name_parts if re.sub(r'[^a-zA-Z]', '', p)]
            
            # Initialize N-Gram vectorizer (2 to 3 character chunks)
            vectorizer = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 3))
            
            for page_num, page in enumerate(self.doc):
                if page_num % 10 == 0: print(".", end="", flush=True)
                
                # Extract all text blocks on the page
                blocks = page.get_text("dict")["blocks"]
                page_texts = []
                page_rects = []
                
                for b in blocks:
                    if "lines" not in b: continue
                    for l in b["lines"]:
                        for s in l["spans"]:
                            text = s["text"].strip().lower()
                            if len(text) > 3: # Ignore tiny noise blocks
                                page_texts.append(text)
                                page_rects.append(fitz.Rect(s["bbox"]))
                
                if not page_texts:
                    continue
                    
                # Vectorize the target name and all page texts together
                # We fit on the page_texts + target to build the vocabulary
                try:
                    tfidf_matrix = vectorizer.fit_transform([clean_target_name] + page_texts)
                except ValueError:
                    continue # Handle empty vocabulary edge cases
                    
                target_vector = tfidf_matrix[0:1]
                page_vectors = tfidf_matrix[1:]
                
                # Calculate N-Gram Cosine Similarity
                cosine_similarities = cosine_similarity(target_vector, page_vectors).flatten()
                
                # Find the best matches on this page
                for idx, n_gram_score in enumerate(cosine_similarities):
                    # Proceed to Phonetic check if N-Gram score shows some promise (> 0.3)
                    if n_gram_score > 0.3:
                        candidate_text = page_texts[idx]
                        candidate_parts = candidate_text.split()
                        
                        # Calculate Phonetic Score
                        phonetic_matches = 0
                        for t_phonetic in target_phonetics:
                            # Clean candidate parts before phonetic matching
                            for c_part in candidate_parts:
                                clean_c = re.sub(r'[^a-zA-Z]', '', c_part)
                                if clean_c and jellyfish.match_rating_codex(clean_c) == t_phonetic:
                                    phonetic_matches += 1
                                    break # Matched this target phonetic once, move to next
                        
                        # Phonetic ratio: how many target parts sounded like the candidate parts?
                        phonetic_score = (phonetic_matches / len(target_phonetics)) if target_phonetics else 0
                        
                        # Fusion Score: 
                        # N-Grams are highly reliable for OCR, Phonetics catch human spelling errors.
                        # If N-Gram is extremely high (>0.85), let it override. Otherwise blend them.
                        if n_gram_score > 0.85:
                            fused_score = n_gram_score
                        else:
                            fused_score = (phonetic_score * 0.5) + (n_gram_score * 0.5)
                        
                        # Convert to 0-100 scale for consistency with old logic
                        fused_score_100 = fused_score * 100

                        
                        if fused_score_100 > best_score:
                            best_score = fused_score_100
                            best_match = {
                                "page": page_num, 
                                "rect": page_rects[idx], 
                                "method": f"hybrid_match ({name}) fused_score: {best_score:.1f}% (N-Gram: {n_gram_score:.2f}, Phonetic: {phonetic_score:.2f})"
                            }
                        
            print(" Done.")

        # Require a solid fused score. 
        # A score > 45 usually means at least one strong phonetic match + decent OCR overlap
        if best_match and best_score > 45:
            return best_match
        
        # 2. Try Plot Number 
        if not title_number:
            return None
            
        numbers = re.findall(r'\d+', str(title_number))
        if not numbers:
             return None
        
        plot_no = numbers[-1] 
        print(f"  Searching for Plot '{plot_no}'...", end="", flush=True)
        
        for page_num, page in enumerate(self.doc):
            # get_text("dict") provides font size, color, and bounding boxes for every text span
            blocks = page.get_text("dict")["blocks"]
            
            candidates = []
            for b in blocks:
                if "lines" not in b: continue
                for l in b["lines"]:
                    for s in l["spans"]:
                        text = s["text"].strip()
                        if not text: continue
                        
                        # Extract digits for comparison
                        clean_num = re.sub(r'[^\d]', '', text)
                        if clean_num != plot_no: continue
                        
                        # DISAMBIGUATION LOGIC:
                        # 1. Ignore if it looks like a dimension (e.g. "39.5", "39m", "39 M")
                        if re.search(r'\d+\.\d+', text): continue # Decimal point = likely dimension
                        if re.search(r'[mM]', text): continue     # 'm' or 'M' suffix = dimension
                        
                        # 2. Strict Numeric Match: 
                        # If we search for "3", we don't want "39" or "13"
                        # re.sub above already gives us the full numeric core of that word
                        if clean_num == plot_no:
                            candidates.append({
                                "rect": fitz.Rect(s["bbox"]),
                                "size": s["size"],
                                "color": s["color"],
                                "text": text
                            })
            
            # Sort candidates by size (Plot numbers are usually larger than dimensions)
            # And color (Black text = 0)
            if candidates:
                # Prioritize: 
                # 1. Exact numeric match with NO extra characters (cleanest hit)
                # 2. Largest font size
                candidates.sort(key=lambda x: (x["text"] == plot_no, x["size"]), reverse=True)
                
                for cand in candidates:
                    rect = cand["rect"]
                    # VICINITY CHECK
                    search_area = rect + (-500, -500, 500, 500)
                    
                    best_v_score = 0
                    best_v_rect = rect
                    best_v_name = ""

                    if name:
                        name_parts = [p for p in name.lower().split() if len(p) > 2]
                        for part in name_parts:
                            part_hits = page.search_for(part, clip=search_area)
                            for p_hit in part_hits:
                                line_box = p_hit + (-200, -10, 200, 10)
                                line_text = page.get_textbox(line_box).replace("\n", " ").strip().lower()
                                
                                words_in_line = line_text.split()
                                matches = 0
                                for p in name_parts:
                                    if any(fuzz.ratio(p, w) > 85 for w in words_in_line):
                                        matches += 1
                                
                                if matches >= 2:
                                    score = fuzz.partial_token_set_ratio(name.lower(), line_text)
                                    if score > best_v_score:
                                        best_v_score = score
                                        best_v_rect = p_hit
                                        best_v_name = name

                    if best_v_score > 75:
                        return {
                            "page": page_num, 
                            "rect": best_v_rect, 
                            "method": f"vicinity_match (Plot {plot_no}, Found {best_v_name} with score {best_v_score})"
                        }
                    
                    # FALLBACK: Return the best numeric hit coordinate
                    return {
                        "page": page_num, 
                        "rect": rect, 
                        "method": f"plot_fallback (Proprietor not found, matched Plot {plot_no} with size {cand['size']:.1f})"
                    }
                    
        return None

    def get_snippet(self, search_result, output_path, aspect_ratio=1.0):
        """
        Renders a RECTANGULAR snippet that matches the target box aspect ratio.
        We use a LARGE base_size to ensure we don't crop out context like names.
        """
        if not search_result:
            return False
            
        page = self.doc[search_result["page"]]
        center = search_result["rect"]
        
        # Sweet spot: 500 units balances legibility and context.
        base_size = 500 
        
        # Adjust dimensions to match the box aspect ratio
        if aspect_ratio >= 1.0: # Landscape/Square target
            crop_w = base_size
            crop_h = base_size / aspect_ratio
        else: # Portrait target
            crop_h = base_size
            crop_w = base_size * aspect_ratio

        clip_rect = fitz.Rect(
            center.x0 - crop_w/2, 
            center.y0 - crop_h/2,
            center.x1 + crop_w/2,
            center.y1 + crop_h/2
        )
        
        # Render
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), clip=clip_rect) 
        pix.save(output_path)
        return True

class PDFProcessor:
    @staticmethod
    def overlay_snippet(consent_pdf_path, snippet_img_path, target_box_1000, output_path, page_index=0, rotation=0):
        # Explicitly cast to int as pandas/numpy types can cause TypeErrors in fitz
        page_index = int(page_index)
        
        doc = fitz.open(consent_pdf_path)
        
        new_doc = fitz.open()
        new_doc.insert_pdf(doc, from_page=page_index, to_page=page_index)
        page = new_doc[0]
        
        # Gemini coordinates are 0-1000 relative to the VISUAL page (what the AI sees).
        # We need to map [ymin, xmin, ymax, xmax] (0-1000) to Internal PDF coordinates.
        
        ymin, xmin, ymax, xmax = target_box_1000
        
        # 1. Map to Visual Points (in PDF points)
        # page.rect is the visual rectangle (width and height already match the rotated view)
        v_w = page.rect.width
        v_h = page.rect.height
        
        v_p0 = fitz.Point(xmin * v_w / 1000, ymin * v_h / 1000)
        v_p1 = fitz.Point(xmax * v_w / 1000, ymax * v_h / 1000)
        
        # INCREASE PADDING: Use 5% padding to create a "centered/floating" look
        padding_x = (v_p1.x - v_p0.x) * 0.05
        padding_y = (v_p1.y - v_p0.y) * 0.05
        v_p0.x += padding_x
        v_p0.y += padding_y
        v_p1.x -= padding_x
        v_p1.y -= padding_y

        # 2. Transform Visual Points back to Internal PDF Space
        # page.rotation_matrix maps internal -> visual. We need visual -> internal.
        mi = ~page.rotation_matrix
        
        i_p0 = v_p0 * mi
        i_p1 = v_p1 * mi
        
        # Create the internal rectangle
        rect = fitz.Rect(i_p0, i_p1)
        rect.normalize()
        
        try:
            # 3. Insert Image
            page.insert_image(rect, filename=snippet_img_path, rotate=rotation)
            new_doc.save(output_path)
            # CRITICAL: Close documents to release file handles
            new_doc.close()
            doc.close()
            return True
        except Exception as e:
            print(f"Error overlaying PDF: {e}")
            if 'new_doc' in locals(): new_doc.close()
            if 'doc' in locals(): doc.close()
            return False

    @staticmethod
    def apply_batch_overlays(source_path, overlay_items, output_path):
        """
        Opens source PDF and applies multiple snippet overlays, then saves a single result.
        overlay_items is a list of dicts: {page_index, snippet_path, box, rotation}
        """
        doc = fitz.open(source_path)
        
        for item in overlay_items:
            try:
                page_index = int(item["page_index"])
                if page_index >= len(doc): continue
                
                page = doc[page_index]
                snippet_path = item["snippet_path"]
                box = item["box"]
                rotation = item.get("rotation", 0)
                
                # Gemini coordinates 0-1000 mapping
                ymin, xmin, ymax, xmax = box
                v_w = page.rect.width
                v_h = page.rect.height
                
                v_p0 = fitz.Point(xmin * v_w / 1000, ymin * v_h / 1000)
                v_p1 = fitz.Point(xmax * v_w / 1000, ymax * v_h / 1000)
                
                # Padding
                padding_x = (v_p1.x - v_p0.x) * 0.05
                padding_y = (v_p1.y - v_p0.y) * 0.05
                v_p0.x += padding_x
                v_p0.y += padding_y
                v_p1.x -= padding_x
                v_p1.y -= padding_y

                mi = ~page.rotation_matrix
                i_p0 = v_p0 * mi
                i_p1 = v_p1 * mi
                rect = fitz.Rect(i_p0, i_p1)
                rect.normalize()
                
                page.insert_image(rect, filename=snippet_path, rotate=rotation)
            except Exception as e:
                print(f"Error in batch overlay on page {item.get('page_index')}: {e}")
                
        try:
            doc.save(output_path)
            doc.close()
            return True
        except Exception as e:
            print(f"Error saving batch PDF: {e}")
            if 'doc' in locals(): doc.close()
            return False

class ExcelWriter:
    @staticmethod
    def _copy_font(f):
        return Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                    vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)

    @staticmethod
    def _copy_border(b):
        return Border(left=copy(b.left), right=copy(b.right), top=copy(b.top),
                      bottom=copy(b.bottom), diagonal=copy(b.diagonal),
                      diagonal_direction=b.diagonal_direction, outline=b.outline,
                      vertical=copy(b.vertical), horizontal=copy(b.horizontal))

    @staticmethod
    def _copy_align(a):
        return Alignment(horizontal=a.horizontal, vertical=a.vertical,
                         text_rotation=a.text_rotation, wrap_text=a.wrap_text,
                         shrink_to_fit=a.shrink_to_fit, indent=a.indent)

    @classmethod
    def _apply_style(cls, src, tgt):
        """Helper to copy style from one cell to another without copying value."""
        if src.has_style:
            tgt.font = cls._copy_font(src.font)
            tgt.border = cls._copy_border(src.border)
            tgt.alignment = cls._copy_align(src.alignment)
            tgt.number_format = src.number_format
            tgt.protection = copy(src.protection)
            tgt.fill = copy(src.fill)

    @classmethod
    def append_data(cls, template_path, data_list, output_buffer):
        """Writes extracted data to Excel, matching row-3 formatting."""
        wb = load_workbook(template_path)
        ws = wb.active

        # Find real last data row (skip empty formatted rows)
        last_row = 1
        for r in range(1, ws.max_row + 1):
            try:
                cell = ws.cell(r, 1)
                # If we hit a merged cell in the first column, we likely hit a footer.
                if type(cell).__name__ == "MergedCell":
                    pass
                elif cell.value is not None:
                    last_row = r
            except Exception:
                pass

        start = last_row + 1
        
        # Capture template formatting before inserting rows
        template_row_index = 3
        template_cells = [copy(c) for c in ws[template_row_index]]

        # Safely insert exactly len(data_list) rows to push existing footers/merged cells downwards
        ws.insert_rows(start, amount=len(data_list))

        # Project Metadata columns (often constant across rows in the template)
        # We inherit these from the template row if the extracted data is empty.
        METADATA_COLS = {2, 3, 4, 5, 6} 

        for i, data in enumerate(data_list):
            row = start + i

            # S/NO
            prev = ws.cell(row - 1, 1).value
            try:
                sno = int(prev) + 1
            except (ValueError, TypeError):
                sno = i + 1
            ws.cell(row, 1, sno)

            # Data columns
            for col, key in COL_MAP.items():
                if col == 1:
                    continue  # S/NO handled above
                
                val = data.get(key, "") if key else ""
                
                # Metadata Inheritance: If extracted value is empty, try to get from template row
                # (This preserves Project Name, County, Region etc. from Row 3)
                if not val and col in METADATA_COLS:
                    val = template_cells[col - 1].value

                # Check for merged cell to be absolutely safe
                target_cell = ws.cell(row, col)
                if type(target_cell).__name__ == "MergedCell":
                    continue
                target_cell.value = val

            # Copy formatting from template row (Columns 1-14)
            for col in range(1, 15):
                src = template_cells[col - 1]
                tgt = ws.cell(row, col)
                cls._apply_style(src, tgt)

        wb.save(output_buffer)
        return len(data_list)
