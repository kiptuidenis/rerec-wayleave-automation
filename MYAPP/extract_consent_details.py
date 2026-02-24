import os
import json
import time
import base64
from pathlib import Path
from copy import copy
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment

# ── Config ──────────────────────────────────────────────────────────────
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")
if not API_KEY:
    raise SystemExit("ERROR: Set GEMINI_API_KEY in your .env file.")

client = genai.Client(api_key=API_KEY)
MODEL = "gemini-3-flash-preview"

IMAGE_DIR = Path("Sample Images")
EXCEL_SRC = Path("SAMITUI VILLAGE  CONSENT LIST.XLSX")
EXCEL_OUT = Path("SAMITUI VILLAGE  CONSENT LIST_UPDATED.XLSX")

MAX_RETRIES = 4
MAX_WORKERS = 4  # parallel image processing

PROMPT = """
Extract the following from this consent form image.
Return ONLY valid JSON -- no markdown, no explanation.

{
  "Project Name": "",
  "Constituency": "",
  "County": "",
  "Plot No": "",
  "Owned by": "",
  "Signed by": "",
  "Relationship": "",
  "ID No": "",
  "Phone No": "",
  "Consent Signed": ""
}

Rules:
- "Project Name" = the name of the project or village on the form.
- "Owned by" = full name of the land owner.
- "Signed by" = full name of the person who signed as proprietor (NOT the witness).
- "Relationship" = relationship of the signer to the owner (e.g. WIFE, HUSBAND, SON, SELF).
- "ID No" = ID number of the proprietor / signer.
- "Phone No" = phone number of the proprietor / signer.
- "Consent Signed" = YES if the proprietor signed the form, otherwise NO.
"""


# ── MIME type lookup ────────────────────────────────────────────────────
def _mime(path: Path) -> str:
    ext = path.suffix.lower()
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png"}.get(ext.lstrip("."), "image/jpeg")


# ── Extraction ──────────────────────────────────────────────────────────
def extract_one(image_path: Path) -> dict | None:
    """Extract consent details from a single image with retries."""
    name = image_path.name

    # Read the image into memory once (avoids repeated file uploads)
    img_bytes = image_path.read_bytes()
    mime = _mime(image_path)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=[
                    types.Part.from_bytes(data=img_bytes, mime_type=mime),
                    PROMPT,
                ],
            )
            text = resp.text

            # Strip markdown fences if present
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]

            data = json.loads(text.strip())
            print(f"  [OK] {name}", flush=True)
            return data

        except Exception as e:
            wait = 2 ** attempt  # 2, 4, 8, 16s
            err_msg = str(e).split("\n")[0][:100]
            print(f"  [FAIL] {name} (attempt {attempt}/{MAX_RETRIES}): {err_msg}", flush=True)
            if attempt < MAX_RETRIES:
                time.sleep(wait)

    print(f"  [GIVE UP] {name} after {MAX_RETRIES} attempts.", flush=True)
    return None


def extract_all(image_paths: list[Path]) -> list[dict]:
    """Process all images in parallel."""
    results = []
    print(f"\nExtracting {len(image_paths)} images ({MAX_WORKERS} parallel workers)...\n", flush=True)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(extract_one, p): p for p in image_paths}
        for future in as_completed(futures):
            data = future.result()
            if data:
                results.append(data)

    return results


# ── Excel helpers ───────────────────────────────────────────────────────
def _copy_font(f):
    return Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)

def _copy_border(b):
    return Border(left=copy(b.left), right=copy(b.right), top=copy(b.top),
                  bottom=copy(b.bottom), diagonal=copy(b.diagonal),
                  diagonal_direction=b.diagonal_direction, outline=b.outline,
                  vertical=copy(b.vertical), horizontal=copy(b.horizontal))

def _copy_align(a):
    return Alignment(horizontal=a.horizontal, vertical=a.vertical,
                     text_rotation=a.text_rotation, wrap_text=a.wrap_text,
                     shrink_to_fit=a.shrink_to_fit, indent=a.indent)


# Column mapping: col_index -> JSON key (None = special handling)
COL_MAP = {
    1: None,                  # S/NO (auto-increment)
    2: "Project Name",
    3: "Constituency",
    4: "County",
    5: None,                  # Region (empty)
    6: None,                  # Affected land (empty)
    7: "Plot No",
    8: "Owned by",
    9: "Signed by",
    10: "Relationship",
    11: "ID No",
    12: "Phone No",
    13: None,                 # Ownership Document (empty)
    14: "Consent Signed",
}


def write_excel(data_list: list[dict]):
    """Write extracted data to Excel, matching row-3 formatting (Goudy Old Style 8pt)."""
    wb = load_workbook(EXCEL_SRC)
    ws = wb.active

    # Find real last data row (skip empty formatted rows)
    last_row = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            last_row = r

    start = last_row + 1
    template = ws[3]  # Row 3 = example data row (Goudy Old Style, 8pt, not bold)

    for i, data in enumerate(data_list):
        row = start + i

        # S/NO
        prev = ws.cell(row - 1, 1).value
        try:
            sno = int(prev) + 1
        except (ValueError, TypeError):
            sno = i + 2
        ws.cell(row, 1, sno)

        # Data columns (including Project Name from col 2)
        for col, key in COL_MAP.items():
            if col == 1:
                continue  # S/NO handled above
            ws.cell(row, col, data.get(key, "") if key else "")

        # Copy formatting from template row
        for col in range(1, 15):
            src = template[col - 1]
            tgt = ws.cell(row, col)
            if src.has_style:
                tgt.font = _copy_font(src.font)
                tgt.border = _copy_border(src.border)
                tgt.alignment = _copy_align(src.alignment)

    wb.save(EXCEL_OUT)
    print(f"\n[DONE] Saved {len(data_list)} rows to {EXCEL_OUT}", flush=True)


# ── Main ────────────────────────────────────────────────────────────────
def main():
    if not IMAGE_DIR.exists():
        raise SystemExit(f"ERROR: '{IMAGE_DIR}' directory not found.")

    images = sorted(p for p in IMAGE_DIR.iterdir()
                    if p.suffix.lower() in ('.jpg', '.jpeg', '.png'))

    if not images:
        raise SystemExit(f"ERROR: No images found in '{IMAGE_DIR}'.")

    print(f"Found {len(images)} images in '{IMAGE_DIR}'.", flush=True)

    t0 = time.time()
    results = extract_all(images)
    elapsed = time.time() - t0

    print(f"\nExtracted {len(results)}/{len(images)} in {elapsed:.1f}s", flush=True)

    if results:
        write_excel(results)
    else:
        print("No data extracted.", flush=True)


if __name__ == "__main__":
    main()
