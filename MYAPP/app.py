import os
import io
import json
import time
import tempfile
from pathlib import Path
from copy import copy
from concurrent.futures import ThreadPoolExecutor, as_completed

import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment

# ── Config ──────────────────────────────────────────────────────────────
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")
MODEL = "gemini-3-flash-preview"
MAX_RETRIES = 4
MAX_WORKERS = 4

EXCEL_TEMPLATE = Path("SAMITUI VILLAGE  CONSENT LIST.XLSX")

COLUMNS = [
    "Project Name", "Constituency", "County", "Plot No",
    "Owned by", "Signed by", "Relationship",
    "ID No", "Phone No", "Consent Signed",
]

PROMPT = """
Extract the following from this consent form image.
Return ONLY valid JSON -- no markdown, no explanation.

{
  "Project Name": "",
  "Constituency": "",
  "County": "",
  "Plot No": "",
  "Owned by": "",
  "Signed by": "",
  "Relationship": "",
  "ID No": "",
  "Phone No": "",
  "Consent Signed": ""
}

Rules:
- "Project Name" = the name of the project or village on the form.
- "Owned by" = full name of the land owner.
- "Signed by" = full name of the person who signed as proprietor (NOT the witness).
- "Relationship" = relationship of the signer to the owner (e.g. WIFE, HUSBAND, SON, SELF).
- "ID No" = ID number of the proprietor / signer.
- "Phone No" = phone number of the proprietor / signer.
- "Consent Signed" = YES if the proprietor signed the form, otherwise NO.
"""

# Column mapping: col_index -> JSON key (None = special handling)
COL_MAP = {
    1: None,              # S/NO (auto-increment)
    2: "Project Name",
    3: "Constituency",
    4: "County",
    5: None,              # Region (empty)
    6: None,              # Affected land (empty)
    7: "Plot No",
    8: "Owned by",
    9: "Signed by",
    10: "Relationship",
    11: "ID No",
    12: "Phone No",
    13: None,             # Ownership Document (empty)
    14: "Consent Signed",
}


# ── Extraction ──────────────────────────────────────────────────────────
def _mime_from_name(name: str) -> str:
    ext = Path(name).suffix.lower().lstrip(".")
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png"}.get(ext, "image/jpeg")


def extract_one_from_bytes(img_bytes: bytes, filename: str, client: genai.Client) -> dict | None:
    """Extract consent details from image bytes with retries."""
    mime = _mime_from_name(filename)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=[
                    types.Part.from_bytes(data=img_bytes, mime_type=mime),
                    PROMPT,
                ],
            )
            text = resp.text

            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]

            return json.loads(text.strip())

        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(2 ** attempt)

    return None


# ── Excel generation ────────────────────────────────────────────────────
def _copy_font(f):
    return Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=f.color)

def _copy_border(b):
    return Border(left=copy(b.left), right=copy(b.right), top=copy(b.top),
                  bottom=copy(b.bottom), diagonal=copy(b.diagonal),
                  diagonal_direction=b.diagonal_direction, outline=b.outline,
                  vertical=copy(b.vertical), horizontal=copy(b.horizontal))

def _copy_align(a):
    return Alignment(horizontal=a.horizontal, vertical=a.vertical,
                     text_rotation=a.text_rotation, wrap_text=a.wrap_text,
                     shrink_to_fit=a.shrink_to_fit, indent=a.indent)


def generate_excel(data_list: list[dict]) -> bytes:
    """Generate Excel file from data list and return as bytes."""
    wb = load_workbook(EXCEL_TEMPLATE)
    ws = wb.active

    # Find real last data row
    last_row = 1
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            last_row = r

    start = last_row + 1
    template = ws[3]  # Example data row (Goudy Old Style, 8pt)

    for i, data in enumerate(data_list):
        row = start + i

        # S/NO
        prev = ws.cell(row - 1, 1).value
        try:
            sno = int(prev) + 1
        except (ValueError, TypeError):
            sno = i + 2
        ws.cell(row, 1, sno)

        # Data columns
        for col, key in COL_MAP.items():
            if col == 1:
                continue
            ws.cell(row, col, data.get(key, "") if key else "")

        # Copy formatting
        for col in range(1, 15):
            src = template[col - 1]
            tgt = ws.cell(row, col)
            if src.has_style:
                tgt.font = _copy_font(src.font)
                tgt.border = _copy_border(src.border)
                tgt.alignment = _copy_align(src.alignment)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ────────────────────────────────────────────────────────
st.set_page_config(page_title="Consent Extractor", page_icon="📋", layout="wide")

# Custom styling
st.markdown("""
<style>
    .main-header {
        font-size: 2rem;
        font-weight: 700;
        background: linear-gradient(90deg, #1a73e8, #4285f4);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        color: #5f6368;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }
    .stDataEditor {
        border: 1px solid #dadce0;
        border-radius: 8px;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">Consent Form Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Upload scanned consent forms, review extracted data, and download the Excel report.</div>', unsafe_allow_html=True)

# ── Check API key ──
if not API_KEY:
    st.error("**GEMINI_API_KEY** not found. Add it to your `.env` file and restart.")
    st.stop()

client = genai.Client(api_key=API_KEY)

# ── Step 1: Upload ──
st.markdown("### 1. Upload Consent Forms")
uploaded_files = st.file_uploader(
    "Select one or more scanned consent form images",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.info(f"**{len(uploaded_files)}** image(s) selected")

# ── Step 2: Extract ──
if uploaded_files and st.button("Extract Details", type="primary", use_container_width=True):
    progress_bar = st.progress(0, text="Starting extraction...")
    status_text = st.empty()

    results = []
    image_map = {}  # filename -> bytes (for viewer)
    total = len(uploaded_files)
    completed = 0
    failed = 0

    # Read all files into memory
    file_data = [(f.read(), f.name) for f in uploaded_files]
    for img_bytes, name in file_data:
        image_map[name] = img_bytes

    def _process(item):
        img_bytes, name = item
        return name, extract_one_from_bytes(img_bytes, name, client)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_process, item): item[1] for item in file_data}

        for future in as_completed(futures):
            name, data = future.result()
            completed += 1

            if data:
                data["Source File"] = name  # track which image this came from
                results.append(data)
                status_text.success(f"[{completed}/{total}] Extracted: {name}")
            else:
                failed += 1
                status_text.warning(f"[{completed}/{total}] Failed: {name}")

            progress_bar.progress(completed / total, text=f"Processing... {completed}/{total}")

    progress_bar.progress(1.0, text="Done!")

    if failed:
        st.warning(f"**{failed}** image(s) failed after {MAX_RETRIES} retries.")

    st.success(f"**{len(results)}/{total}** consent forms extracted successfully.")

    # Store in session state
    st.session_state["extracted_data"] = results
    st.session_state["image_map"] = image_map

# ── Step 3: Editable table + Image viewer ──
if "extracted_data" in st.session_state and st.session_state["extracted_data"]:
    st.markdown("### 2. Review & Edit Extracted Data")
    st.caption("Edit any cell in the table. Select a source file from the dropdown to view the original image.")

    display_cols = ["Source File"] + COLUMNS
    df = pd.DataFrame(st.session_state["extracted_data"], columns=display_cols)

    # Add S/NO column
    df.insert(0, "S/NO", range(1, len(df) + 1))

    # Layout: table on the left, image viewer on the right
    table_col, viewer_col = st.columns([3, 2])

    with table_col:
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "S/NO": st.column_config.NumberColumn("S/NO", disabled=True, width="small"),
                "Source File": st.column_config.TextColumn("Source File", disabled=True, width="medium"),
                "Consent Signed": st.column_config.SelectboxColumn(
                    "Consent Signed",
                    options=["YES", "NO"],
                    width="small",
                ),
            },
        )

    with viewer_col:
        st.markdown("#### Image Viewer")
        image_map = st.session_state.get("image_map", {})
        if image_map:
            filenames = list(image_map.keys())
            selected = st.selectbox("Select image to view", filenames, label_visibility="collapsed")
            if selected and selected in image_map:
                st.image(image_map[selected], caption=selected, use_container_width=True)
        else:
            st.info("No images available for preview.")

    # ── Step 4: Download ──
    st.markdown("### 3. Download Excel")

    # Convert edited dataframe back to list of dicts (drop S/NO and Source File — not in Excel)
    export_data = edited_df.drop(columns=["S/NO", "Source File"], errors="ignore").to_dict(orient="records")

    if EXCEL_TEMPLATE.exists():
        excel_bytes = generate_excel(export_data)
        st.download_button(
            label="Download Updated Excel",
            data=excel_bytes,
            file_name="SAMITUI VILLAGE CONSENT LIST_UPDATED.XLSX",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    else:
        st.warning(f"Template file `{EXCEL_TEMPLATE}` not found. Place it in the app directory.")
